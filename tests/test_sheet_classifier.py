import os
import datetime
import openpyxl
import pytest
from engine.config import ClientProfile, kane_jones_profile
from engine.sheet_classifier import classify_workbook_sheets, ClassificationReport
from engine.parser import _looks_like_date, _coerce_date, parse_workbook


@pytest.fixture
def sample_xlsx_path():
    candidates = [
        os.path.join("sample_data", "July_sales_report_v6.xlsx"),
        os.path.join("sample_data", "July_sales_report_v7 - Copy.xlsx"),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    pytest.skip("Benchmark July file not found in sample_data/")



def test_classify_standard_workbook(sample_xlsx_path):
    profile = kane_jones_profile()
    report: ClassificationReport = classify_workbook_sheets(sample_xlsx_path, profile)

    # 1. Sales sheets (both hierarchical raw dumps identified)
    assert "tmpA1A6" in report.sales_sheets
    assert "tmp32C7" in report.sales_sheets
    assert len(report.sales_sheets) == 2

    # 2. Inventory sheet
    assert report.inventory_sheet == "tmp3F5D"

    # 3. Sales returns sheet
    assert report.sales_returns_sheet == "tmpCEF3"

    # 4. Purchases and Purchase Returns sheets
    assert report.purchases_sheet == "tmp6434"
    assert report.purchase_returns_sheet == "tmpBAA5"
    assert report.stock_balances_sheet == "tnp7rq1"



def test_classify_dedicated_price_list(tmp_path):
    """Test dedicated 3-tier price list detection."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Price Matrix"
    ws.append(["SKU Description", "Distributor Price", "Sub-Distributor Price", "Retail Price"])
    ws.append(["Heineken 33cl", 12000, 12500, 13000])
    ws.append(["Goldberg 60cl", 8500, 8800, 9200])

    path = str(tmp_path / "pricing.xlsx")
    wb.save(path)

    profile = ClientProfile(client_id="test", display_name="Test Depot", price_list_sheet="")
    report = classify_workbook_sheets(path, profile)

    assert report.price_list_sheet == "Master Price Matrix"


def test_classify_renamed_sheets(sample_xlsx_path, tmp_path):
    """Stress test: sheets renamed to arbitrary names without profile hints."""
    wb = openpyxl.load_workbook(sample_xlsx_path)
    wb["tmpA1A6"].title = "tmpZZ99_Sales"
    wb["tmp32C7"].title = "tmpYY88_Sales"
    wb["tmp3F5D"].title = "tmpINV77_Stock"
    wb["tmpCEF3"].title = "tmpRET66_Returns"

    renamed_path = str(tmp_path / "renamed.xlsx")
    wb.save(renamed_path)

    # Generic profile with no sheet hints
    generic_profile = ClientProfile(
        client_id="generic",
        display_name="Generic Depot",
        raw_data_sheets=[],
        price_list_sheet="Price list",
        inventory_sheet="",
        sales_returns_sheet="",
        expenses_sheet="",
    )

    report = classify_workbook_sheets(renamed_path, generic_profile)

    assert "tmpZZ99_Sales" in report.sales_sheets
    assert "tmpYY88_Sales" in report.sales_sheets
    assert report.inventory_sheet == "tmpINV77_Stock"
    assert report.sales_returns_sheet == "tmpRET66_Returns"


# ── Fix B1: String date recognition ────────────────────────────────────────────

def test_looks_like_date_native_types():
    """Native datetime and date objects are always accepted."""
    assert _looks_like_date(datetime.datetime(2026, 7, 1, 10, 30))
    assert _looks_like_date(datetime.date(2026, 7, 1))


def test_looks_like_date_string_formats():
    """Common ERP-exported string date formats are accepted."""
    assert _looks_like_date("01/07/2026")   # DD/MM/YYYY
    assert _looks_like_date("2026-07-01")   # ISO
    assert _looks_like_date("01-07-2026")   # DD-MM-YYYY
    assert _looks_like_date("2026/07/01")   # YYYY/MM/DD


def test_looks_like_date_rejects_non_dates():
    """Non-date strings and other types are rejected."""
    assert not _looks_like_date("ITEM/SERVICE")
    assert not _looks_like_date("Heineken 60cl")
    assert not _looks_like_date(12345.0)
    assert not _looks_like_date("")
    assert not _looks_like_date(None)


def test_coerce_date_string():
    """_coerce_date correctly converts string dates to datetime objects."""
    dt = _coerce_date("01/07/2026")
    assert isinstance(dt, datetime.datetime)
    assert dt.year == 2026 and dt.month == 7 and dt.day == 1

    dt2 = _coerce_date("2026-08-15")
    assert dt2.year == 2026 and dt2.month == 8 and dt2.day == 15


# ── Fix B2: Fully-unknown sheet names parse end-to-end ─────────────────────────

def test_parse_workbook_with_renamed_sheets_extracts_invoices(sample_xlsx_path, tmp_path):
    """Regression test for Bug B2 / the tmp6D3B scenario.

    Simulates exactly what happened in production: the ERP export regenerated
    sheet names (tmpA1A6 → tmpXX99, tmp32C7 → tmpYY88). The profile whitelist
    no longer matches. The structural classifier must find the sheets and the
    parser must extract all invoices without any literal-name gate.
    """
    wb = openpyxl.load_workbook(sample_xlsx_path)
    wb["tmpA1A6"].title = "tmpXX99"
    wb["tmp32C7"].title = "tmpYY88"

    renamed_path = str(tmp_path / "renamed_sales.xlsx")
    wb.save(renamed_path)

    # Profile still carries the old stale names — must NOT be used to gate parsing
    profile = kane_jones_profile()
    assert profile.raw_data_sheets == ["tmpA1A6", "tmp32C7"], "Precondition: profile still has old names"

    inv_df, li_df, anom_df = parse_workbook(renamed_path, profile)

    # Must extract the same number of invoices as the original workbook
    original_inv_df, _, _ = parse_workbook(sample_xlsx_path, profile)

    assert not inv_df.empty, "No invoices extracted — structural classifier failed to find renamed sheets"
    assert len(inv_df) == len(original_inv_df), (
        f"Renamed workbook: {len(inv_df)} invoices; original: {len(original_inv_df)} invoices — mismatch"
    )
    assert float(inv_df["gross_revenue"].sum()) == pytest.approx(
        float(original_inv_df["gross_revenue"].sum()), rel=1e-4
    ), "Gross revenue totals differ after rename — calculation regression"


# ── Full Automated Ledger Sheets Classification & Parsing Tests ──────────────

def test_classify_and_parse_all_ledger_sheets(sample_xlsx_path):
    """Verifies that Stock Balances, Purchases, and Purchase Returns are classified and parsed."""
    from engine.parser import (
        parse_stock_balances_sheet,
        parse_purchases_sheet,
        parse_purchase_returns_sheet,
    )
    profile = kane_jones_profile()
    rep = classify_workbook_sheets(sample_xlsx_path, profile)

    # July sample has Stock Balances (tnp7rq1), Purchases (tmp6434), Purchase Returns (tmpBAA5)
    assert rep.stock_balances_sheet is not None
    assert rep.purchases_sheet is not None
    assert rep.purchase_returns_sheet is not None

    open_stock, close_stock = parse_stock_balances_sheet(sample_xlsx_path, profile, classification_report=rep)
    assert open_stock == 455005282.0
    assert close_stock == 319967133.0

    total_pur, df_pur, _ = parse_purchases_sheet(sample_xlsx_path, profile, classification_report=rep)
    assert total_pur == 77081290.0
    assert len(df_pur) > 0

    total_pr, df_pr, _ = parse_purchase_returns_sheet(sample_xlsx_path, profile, classification_report=rep)
    assert total_pr == 38589715.0
    assert len(df_pr) > 0

