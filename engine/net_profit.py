"""
Net Profit / Loss Bridge Module for Kane-Jones Depot.

Computes the official management bridge:
  Gross Sales Revenue (all invoices, incl. empties)
- Total Sales Returns (all credit notes, incl. empties)
= Net Sales Revenue
- Net Product COGS (Invoiced product cost - Cost of returned goods credited back)
= Net Gross Profit / (Loss)
- Total Operating Expenses
= Net Operating Profit / (Loss)

Matches the exact structure of the 'Sales Return Analysis' reference sheet.
"""

from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
import openpyxl
from engine.config import ClientProfile


def _parse_numeric(val: Any) -> Optional[float]:
    """Helper to cleanly parse numbers from formatted strings (e.g. '2,095,229', '₦10,000.00')."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("₦", "").replace("NGN", "").replace("$", "").replace("%", "").strip()
    try:
        return float(s)
    except Exception:
        return None


def parse_expenses_sheet(
    xlsx_path_or_wb: Any,
    profile: ClientProfile = None,
    classification_report: Optional[Any] = None,
) -> Tuple[float, pd.DataFrame, List[Dict[str, Any]]]:
    """
    Parses operating expenses from an expense workbook (e.g. july_expn.xlsx)
    or an in-workbook expenses sheet (e.g. 'July total Expenses', 'July Expenses threshold ', tmp6F17).
    
    Supports:
    1. Category / Summary Tables (Col 1 = Category, Col 2 = Amount, Grand Total row).
    2. Day-Book Ledger Format (Date, Category, Payment, Voucher Number, Amount).
    Returns: (total_expenses, categories_df, anomalies)
    """
    if profile is None:
        from engine.config import kane_jones_profile
        profile = kane_jones_profile()

    if isinstance(xlsx_path_or_wb, str):
        try:
            wb = openpyxl.load_workbook(xlsx_path_or_wb, data_only=True)
        except Exception:
            return 0.0, pd.DataFrame(), []
    elif hasattr(xlsx_path_or_wb, "sheetnames"):
        wb = xlsx_path_or_wb
    else:
        return 0.0, pd.DataFrame(), []

    anomalies = []
    candidate_sheets = []
    
    # 0. Check classification report if provided
    if classification_report is not None:
        exp_sheet = getattr(classification_report, "expenses_sheet", None)
        if exp_sheet and exp_sheet in wb.sheetnames:
            candidate_sheets.append(exp_sheet)
        # If classifier found no expenses sheet, still try the profile-named sheet
        # before falling through to heuristic name scanning below.
        # DO NOT bail out early here — expenses sheet may be present but misclassified.
        elif getattr(profile, "expenses_sheet", None) and profile.expenses_sheet in wb.sheetnames:
            candidate_sheets.append(profile.expenses_sheet)

    # 1. Identify Candidate Sheets (Prioritizing clean expense summary sheets over ledgers)
    summary_candidates = []
    other_candidates = []
    
    # Exclude sheets known to be non-expenses
    excluded_sheets = set()
    if classification_report is not None:
        for s_name, c_info in getattr(classification_report, "classifications", {}).items():
            if c_info.role in ("sales_invoice", "inventory_cost", "sales_returns", "price_list"):
                excluded_sheets.add(s_name)
        for s_name in getattr(classification_report, "sales_sheets", []):
            excluded_sheets.add(s_name)
    
    for s in wb.sheetnames:
        if s in candidate_sheets or s in excluded_sheets:
            continue
        clean = s.strip().lower()
        if any(skw in clean for skw in ["sales", "revenue", "price", "inventory", "aggregate", "customer", "marketer", "product", "credit", "return", "cash", "gtb"]):
            continue
        if any(kw in clean for kw in ["expense", "expn", "exp", "opex", "overhead", "spending"]):
            if any(kw in clean for kw in ["threshold", "summary", "total", "all", "cat"]):
                summary_candidates.append(s)
            else:
                other_candidates.append(s)
        elif any(kw in clean for kw in ["voucher", "payment", "6f17"]):
            other_candidates.append(s)

    if not candidate_sheets:
        candidate_sheets = summary_candidates + other_candidates
    else:
        for s in summary_candidates + other_candidates:
            if s not in candidate_sheets:
                candidate_sheets.append(s)

    # 2. Attempt parsing candidate sheets (Summary Table format first, then Day-book format)
    for sheet_name in candidate_sheets:
        ws = wb[sheet_name]
        
        # Check if first row/header looks like DayBook (VOUCHER DATE, DEBIT, CREDIT)
        is_daybook = False
        for r in range(1, min(ws.max_row or 1, 6) + 1):
            row_vals = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, min(ws.max_column or 1, 8) + 1)]
            if "DAY BOOK" in row_vals or ("DEBIT" in row_vals and "CREDIT" in row_vals and "VCH TYPE" in row_vals):
                is_daybook = True
                break
                
        if not is_daybook:
            # Test Format A: Category / Summary Table (e.g. 'July total Expenses', 'July Expenses threshold ')
            categories = []
            grand_total = 0.0
            is_summary_format = False
            
            for r in range(1, (ws.max_row or 1) + 1):
                c1 = ws.cell(r, 1).value
                c2 = ws.cell(r, 2).value
                if c1 is None and c2 is None:
                    continue
                str_c1 = str(c1 or "").strip()
                str_c2 = str(c2 or "").strip()
                
                # Header check
                if "category" in str_c1.lower() and ("amount" in str_c2.lower() or "%" in str_c2.lower() or "cost" in str_c2.lower()):
                    is_summary_format = True
                    continue
                    
                val2 = _parse_numeric(c2)
                # Grand Total check
                if "grand total" in str_c1.lower() or "total expense" in str_c1.lower() or (str_c1.lower() == "total" and val2 is not None):
                    is_summary_format = True
                    if val2 is not None and val2 > 0:
                        grand_total = val2
                    continue
                    
                # Category line item
                if str_c1 and val2 is not None and val2 > 0 and not str_c1.startswith("---"):
                    # Avoid day-book header / date rows
                    if not str_c1.lower().startswith("voucher") and not str_c1.lower().startswith("date") and not str_c1.lower().startswith("vch"):
                        categories.append({
                            "category": str_c1,
                            "amount": val2,
                            "source_row": r,
                        })
                        
            if categories and (is_summary_format or len(categories) >= 2):
                df_exp = pd.DataFrame(categories)
                if grand_total == 0.0 and not df_exp.empty:
                    grand_total = float(df_exp["amount"].sum())
                return grand_total, df_exp, anomalies

        # Test Format B: Day-Book Ledger Format (e.g. 'tmp6F17')
        if is_daybook:
            daybook_expenses = []
            daybook_total = 0.0
            
            for r in range(1, (ws.max_row or 1) + 1):
                c1 = ws.cell(r, 1).value
                c2 = ws.cell(r, 2).value
                c4 = ws.cell(r, 4).value
                c5 = ws.cell(r, 5).value
                c6 = ws.cell(r, 6).value

                if c5 and "TOTAL" in str(c5).upper():
                    total_parsed = _parse_numeric(c6)
                    if total_parsed is not None:
                        daybook_total = total_parsed
                    continue

                if c6 is not None:
                    amt = _parse_numeric(c6)
                    c4_str = str(c4).strip().lower() if c4 else ""
                    c2_str = str(c2).strip() if c2 else "General"
                    if amt is not None and amt > 0:
                        if c4_str in ("payment", "journal") or "journal" in c2_str.lower():
                            daybook_expenses.append({
                                "date": str(c1)[:10] if c1 else "",
                                "category": c2_str,
                                "voucher_no": str(c5).strip() if c5 else "",
                                "amount": amt,
                                "source_row": r
                            })

            if daybook_expenses:
                df_daybook_raw = pd.DataFrame(daybook_expenses)
                # Aggregate by unique category name to prevent duplication in category tables
                df_daybook = df_daybook_raw.groupby("category", as_index=False).agg(
                    amount=("amount", "sum"),
                    voucher_count=("voucher_no", "count"),
                    source_row=("source_row", "first")
                ).sort_values(by="amount", ascending=False).reset_index(drop=True)
                
                if daybook_total == 0.0 and not df_daybook.empty:
                    daybook_total = float(df_daybook["amount"].sum())
                return daybook_total, df_daybook, anomalies

    return 0.0, pd.DataFrame(), anomalies



def calculate_financial_statements(
    gross_sales: float,
    sales_returns: float = 0.0,
    purchases: float = 0.0,
    purchase_returns: float = 0.0,
    carriage_inwards: float = 0.0,
    opening_inventory: float = 0.0,
    closing_inventory: float = 0.0,
    operating_expenses: Optional[List[float]] = None,
    other_income: float = 0.0,
    finance_costs: float = 0.0,
) -> Dict[str, float]:
    """
    Standard accounting P&L calculation structure:
      net_sales = gross_sales - sales_returns
      net_purchases = purchases - purchase_returns + carriage_inwards
      cogs = opening_inventory + net_purchases - closing_inventory
      gross_profit = net_sales - cogs
      gross_margin_pct = (gross_profit / net_sales) * 100 if net_sales > 0 else 0.0
      total_expenses = sum(operating_expenses)
      net_profit = (gross_profit + other_income) - total_expenses - finance_costs
      net_margin_pct = (net_profit / net_sales) * 100 if net_sales > 0 else 0.0
    """
    net_sales = float(gross_sales) - float(sales_returns)
    net_purchases = float(purchases) - float(purchase_returns) + float(carriage_inwards)
    cogs = float(opening_inventory) + net_purchases - float(closing_inventory)
    gross_profit = net_sales - cogs
    gross_margin_pct = (gross_profit / net_sales * 100.0) if net_sales > 0 else 0.0
    total_expenses = sum(operating_expenses) if operating_expenses else 0.0
    net_profit = (gross_profit + float(other_income)) - float(total_expenses) - float(finance_costs)
    net_margin_pct = (net_profit / net_sales * 100.0) if net_sales > 0 else 0.0

    return {
        "gross_sales": float(gross_sales),
        "sales_returns": float(sales_returns),
        "net_sales": net_sales,
        "purchases": float(purchases),
        "purchase_returns": float(purchase_returns),
        "carriage_inwards": float(carriage_inwards),
        "net_purchases": net_purchases,
        "opening_inventory": float(opening_inventory),
        "closing_inventory": float(closing_inventory),
        "cogs": cogs,
        "gross_profit": gross_profit,
        "gross_margin_pct": gross_margin_pct,
        "total_expenses": float(total_expenses),
        "other_income": float(other_income),
        "finance_costs": float(finance_costs),
        "net_profit": net_profit,
        "net_margin_pct": net_margin_pct,
    }


def compute_net_profit_bridge(
    invoices_df: pd.DataFrame,
    line_items_df: pd.DataFrame,
    df_returns: pd.DataFrame,
    expenses_total: float = 0.0,
    df_inv: Optional[pd.DataFrame] = None,
    cost_of_returns: Optional[float] = None,
    # The seven accounting inputs below use None as a sentinel to distinguish
    # "caller explicitly passed 0.0" from "caller did not supply this value at all".
    # When None, the field is tracked in missing_accounting_fields and defaults to 0.0.
    purchases: Optional[float] = None,
    purchase_returns: Optional[float] = None,
    carriage_inwards: Optional[float] = None,
    carriage_outwards: Optional[float] = None,
    opening_inventory: Optional[float] = None,
    closing_inventory: Optional[float] = None,
    other_income: Optional[float] = None,
    finance_costs: Optional[float] = None,
    operating_expenses: Optional[List[float]] = None,
    profile: ClientProfile = None
) -> Dict[str, Any]:
    """
    Computes the complete Net Profit / Loss bridge from raw transaction data.

    Accounting Formula:
      net_sales = gross_sales - sales_returns
      net_purchases = purchases - purchase_returns + carriage_inwards
      cogs = opening_inventory + net_purchases - closing_inventory
      gross_profit = net_sales - cogs
      gross_margin_pct = (gross_profit / net_sales) * 100 if net_sales > 0 else 0.0
      total_expenses = sum(operating_expenses)
      net_profit = (gross_profit + other_income) - total_expenses - finance_costs
      net_margin_pct = (net_profit / net_sales) * 100 if net_sales > 0 else 0.0

    Data Integrity Rules:
      - sales_returns NEVER deducted from cost; only reduces net_sales.
      - purchase_returns only reduces net_purchases (COGS).
      - cost_of_returns is maintained for audit transparency but NOT deducted from COGS.

    Missing Field Warnings:
      Any of the seven ledger inputs not supplied by the caller will appear in
      missing_accounting_fields in the return dict. Callers MUST surface this list
      to end users so they know which figures were computed with 0.0 assumptions.
    """
    if profile is None:
        from engine.config import kane_jones_profile
        profile = kane_jones_profile()

    # Detect which of the 7 accounting inputs were not supplied by the caller.
    # None sentinel = "not provided"; 0.0 = "explicitly zeroed by caller".
    _SEVEN_FIELDS = {
        "purchases": purchases,
        "purchase_returns": purchase_returns,
        "carriage_inwards": carriage_inwards,
        "opening_inventory": opening_inventory,
        "closing_inventory": closing_inventory,
        "other_income": other_income,
        "finance_costs": finance_costs,
    }
    missing_accounting_fields = [k for k, v in _SEVEN_FIELDS.items() if v is None]

    # Resolve None → 0.0 for calculation (purchases falls back to embedded cost later)
    purchase_returns_val = float(purchase_returns) if purchase_returns is not None else 0.0
    carriage_inwards_val = float(carriage_inwards) if carriage_inwards is not None else 0.0
    carriage_outwards_val = float(carriage_outwards) if carriage_outwards is not None else 0.0
    opening_inventory_val = float(opening_inventory) if opening_inventory is not None else 0.0
    closing_inventory_val = float(closing_inventory) if closing_inventory is not None else 0.0
    other_income_val = float(other_income) if other_income is not None else 0.0
    finance_costs_val = float(finance_costs) if finance_costs is not None else 0.0

    if invoices_df is not None and not invoices_df.empty and "gross_revenue" in invoices_df.columns:
        gross_sales_revenue = float(invoices_df["gross_revenue"].sum())
    elif not line_items_df.empty:
        gross_sales_revenue = float(line_items_df["quantity"].mul(line_items_df["rate"]).sum())
    else:
        gross_sales_revenue = 0.0

    # Full Invoice-Embedded Cost (all lines including empties, matching invoice header total cost)
    if not line_items_df.empty and "cost" in line_items_df.columns:
        gross_embedded_cost = float(line_items_df["cost"].sum())
    elif invoices_df is not None and not invoices_df.empty and "invoice_cost" in invoices_df.columns:
        gross_embedded_cost = float(invoices_df["invoice_cost"].sum())
    else:
        gross_embedded_cost = 0.0

    # Sales returns breakdown
    total_sales_returns = 0.0
    product_returns_val = 0.0
    product_returns_qty = 0.0
    empties_returns_val = 0.0
    empties_returns_qty = 0.0

    if df_returns is not None and not df_returns.empty:
        total_sales_returns = float(df_returns["return_value"].sum())
        prod_ret = df_returns[df_returns["item_type"] == "Product"]
        emp_ret = df_returns[df_returns["item_type"] == "Empties"]
        product_returns_val = float(prod_ret["return_value"].sum()) if not prod_ret.empty else 0.0
        product_returns_qty = float(prod_ret["quantity"].sum()) if not prod_ret.empty else 0.0
        empties_returns_val = float(emp_ret["return_value"].sum()) if not emp_ret.empty else 0.0
        empties_returns_qty = float(emp_ret["quantity"].sum()) if not emp_ret.empty else 0.0

    # Calculate Cost of Returns Credited Back (audit trail only)
    missing_cost_anomalies = []
    if cost_of_returns is None:
        cost_of_returns = 0.0
        if df_returns is not None and not df_returns.empty:
            from engine.true_cost import resolve_sku_cost_maps, resolve_return_unit_cost
            unit_cost_map = resolve_sku_cost_maps(line_items_df, df_inv, prefer_inventory=False)

            for _, r in df_returns.iterrows():
                item_name = str(r.get("item_name", "")).strip()
                unit_c, anom = resolve_return_unit_cost(item_name, unit_cost_map)
                qty = float(r.get("quantity", 0.0) or 0.0)
                if unit_c is not None:
                    cost_of_returns += (qty * unit_c)
                elif anom is not None:
                    anom["voucher_no"] = r.get("voucher_no")
                    anom["customer"] = r.get("customer")
                    anom["quantity"] = qty
                    anom["return_value"] = float(r.get("return_value", 0.0) or 0.0)
                    missing_cost_anomalies.append(anom)

    # 1. Net Sales
    net_sales_revenue = gross_sales_revenue - total_sales_returns

    # 2. Net Purchases & COGS
    purchases_val = float(purchases) if purchases is not None else gross_embedded_cost
    net_purchases = purchases_val - purchase_returns_val + carriage_inwards_val
    cogs = opening_inventory_val + net_purchases - closing_inventory_val
    net_cost = cogs

    # 3. Gross Profit
    gross_profit = net_sales_revenue - cogs
    net_gross_margin_pct = (gross_profit / net_sales_revenue) if net_sales_revenue > 0 else 0.0
    return_rate = (total_sales_returns / gross_sales_revenue) if gross_sales_revenue > 0 else 0.0

    # 4. Operating Expenses & Net Operating Profit
    if operating_expenses is not None:
        tot_expenses = float(sum(operating_expenses))
    else:
        tot_expenses = float(expenses_total or 0.0)
    if carriage_outwards_val:
        tot_expenses += carriage_outwards_val

    net_profit = (gross_profit + other_income_val) - tot_expenses - finance_costs_val
    net_operating_margin_pct = (net_profit / net_sales_revenue) if net_sales_revenue > 0 else 0.0

    return {
        "gross_sales_revenue": gross_sales_revenue,
        "total_sales_returns": total_sales_returns,
        "net_sales_revenue": net_sales_revenue,
        "gross_product_cost": gross_embedded_cost,
        "gross_embedded_cost": gross_embedded_cost,
        "purchases": purchases_val,
        "purchase_returns": purchase_returns_val,
        "carriage_inwards": carriage_inwards_val,
        "carriage_outwards": carriage_outwards_val,
        "opening_inventory": opening_inventory_val,
        "closing_inventory": closing_inventory_val,
        "net_purchases": net_purchases,
        "cogs": cogs,
        "total_cost": cogs,
        "net_cost": net_cost,
        "total_cost_embedded": net_cost,
        "cost_of_returns": float(cost_of_returns or 0.0),
        "gross_profit": gross_profit,
        "net_gross_profit_loss": gross_profit,
        "gross_margin_pct": net_gross_margin_pct * 100.0,
        "net_gross_margin_pct": net_gross_margin_pct,
        "total_operating_expenses": tot_expenses,
        "other_income": other_income_val,
        "finance_costs": finance_costs_val,
        "net_profit": net_profit,
        "net_operating_profit_loss": net_profit,
        "net_operating_margin_pct": net_operating_margin_pct,
        "net_margin_pct": net_operating_margin_pct * 100.0,
        "product_returns_value": product_returns_val,
        "product_returns_qty": product_returns_qty,
        "empties_returns_value": empties_returns_val,
        "empties_returns_qty": empties_returns_qty,
        "return_rate": return_rate,
        "missing_cost_anomalies": missing_cost_anomalies,
        # Fix A: Explicit list of the 7 ledger inputs that were not supplied by the caller.
        # When non-empty, these figures were computed with 0.0 assumptions and MUST be
        # shown as incomplete in the UI / CLI output rather than presented as final.
        "missing_accounting_fields": missing_accounting_fields,
    }


