"""
Matches messy product names as they appear in the sales register (e.g. "33 Export 60cl",
"Goldberg 60cl") against the official price list's SKU names (e.g. "33 Export Bottle 60c1",
"Goldberg Bottle 60c1"). Register names are shorthand/typo'd versions of SKU names — this
is the layer that made "typo-normalized" / "resolved by price tier" matching possible in
the original report.

Strategy:
1. Exact match (case/space-insensitive) first — cheapest and safest.
2. Manual overrides check (profile.manual_overrides) for unambiguous shorthand names.
3. Extract normalized pack size token (cl, ml, L, cases) from product and SKUs.
4. If pack size is present, filter match candidates to SKUs sharing that exact pack size.
5. Re-run fuzzy text matching only within that size-filtered candidate set.
6. If no SKU shares the pack size, fall back to fuzzy match on base name and flag
   match_method as "fuzzy_no_size_match" (lower confidence).
7. Reject matches below profile.fuzzy_match_threshold — better to flag "unmatched" than
   silently attach the wrong price to a product.
"""

import re
from typing import Optional
import pandas as pd
from rapidfuzz import fuzz, process
from engine.config import ClientProfile


def extract_pack_size(name: str) -> Optional[str]:
    """Extract a normalized pack size token from a product name or SKU.
    Handles units: cl, c1 (typo), ml, l, ltr, 1tr/Itr (typos), cases, packs.
    Returns normalized strings like '60CL', '330ML', '1L', '20L', '24PK', '12CASE'.
    """
    if not name or pd.isna(name):
        return None
    pattern = r'(?i)\b(\d+(?:\.\d+)?)\s*(cl|c1|c[iI]|ml|m1|m[iI]|ltr|1tr|[iI]tr|litres?|liters?|l|pk|pack|packs|case|cases|crates?)\b'
    m = re.search(pattern, str(name))
    if not m:
        return None
    val_str, unit = m.group(1), m.group(2).upper()
    val = float(val_str)
    val_norm = int(val) if val.is_integer() else val

    if unit in ('CL', 'C1', 'CI'):
        return f'{val_norm}CL'
    elif unit in ('ML', 'M1', 'MI'):
        return f'{val_norm}ML'
    elif unit in ('L', 'LTR', '1TR', 'ITR', 'LITRE', 'LITRES', 'LITER', 'LITERS'):
        return f'{val_norm}L'
    elif unit in ('PK', 'PACK', 'PACKS'):
        return f'{val_norm}PK'
    elif unit in ('CASE', 'CASES', 'CRATE', 'CRATES'):
        return f'{val_norm}CASE'
    return f'{val_norm}{unit}'


def extract_base_name(name: str) -> str:
    """Extracts product base name by stripping pack size token and extra punctuation."""
    if not name or pd.isna(name):
        return ''
    pattern = r'(?i)\b\d+(?:\.\d+)?\s*(cl|c1|c[iI]|ml|m1|m[iI]|ltr|1tr|[iI]tr|litres?|liters?|l|pk|pack|packs|case|cases|crates?)\b'
    cleaned = re.sub(pattern, ' ', str(name))
    cleaned = re.sub(r'[\.\,\-\_\/]', ' ', cleaned)
    return re.sub(r'\s+', ' ', cleaned).strip()


def normalize_text(text: str) -> str:
    """Normalizes OCR/typo variants in text like '60c1' -> '60cl' and '201tr' -> '20ltr'."""
    if not text or pd.isna(text):
        return ''
    t = re.sub(r'(?i)\b(\d+(?:\.\d+)?)\s*c1\b', r'\g<1>cl', str(text))
    t = re.sub(r'(?i)\b(\d+(?:\.\d+)?)\s*(?:1tr|Itr)\b', r'\g<1>ltr', t)
    return re.sub(r'\s+', ' ', t).strip()


def load_price_list(xlsx_path: str, profile: ClientProfile) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    target_sheet = None
    header_idx = None
    
    # 1. First, check if exact profile.price_list_sheet exists
    candidate_sheets = []
    if profile.price_list_sheet in wb.sheetnames:
        candidate_sheets.append(profile.price_list_sheet)
    
    # 2. Check normalized sheet names
    norm_expected = re.sub(r'[\s_\-\.]', '', profile.price_list_sheet).upper()
    for name in wb.sheetnames:
        if re.sub(r'[\s_\-\.]', '', name).upper() == norm_expected and name not in candidate_sheets:
            candidate_sheets.append(name)
            
    # 3. Add other sheets containing "price" or "pricing"
    for name in wb.sheetnames:
        if ("price" in name.lower() or "pricing" in name.lower()) and name not in candidate_sheets:
            candidate_sheets.append(name)

    # 4. Search candidate sheets for a valid header row containing "SKU" or "DESCRIPTION"
    for s_name in candidate_sheets:
        ws = wb[s_name]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        for i, row in enumerate(rows):
            joined = " ".join(str(c) for c in row if c).upper()
            if ("SKU" in joined or "DESCRIPTION" in joined) and ("DISTRIBUTOR" in joined or "PRICE" in joined or "RATE" in joined or "TIER" in joined):
                target_sheet = s_name
                header_idx = i
                break
        if target_sheet is not None:
            break

    # 5. If a valid price list sheet was found with standard header:
    if target_sheet is not None and header_idx is not None:
        ws = wb[target_sheet]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        cols = ["sku", "distributor_price", "sub_distributor_price", "retail_price"]
        data_rows = [r[:len(cols)] for r in rows[header_idx + 1:] if r[0] not in (None, "")]
        price_list = pd.DataFrame(data_rows, columns=cols)
        price_list = price_list.dropna(subset=["sku"]).reset_index(drop=True)
        for col in ["distributor_price", "sub_distributor_price", "retail_price"]:
            if col in price_list.columns:
                price_list[col] = pd.to_numeric(price_list[col], errors="coerce")
        return price_list

    # 6. Fallback: If no dedicated 3-tier price list sheet exists, check if inventory sheet (tmp3F5D) exists
    from engine.parser import parse_inventory_sheet
    inv_sheet_name = getattr(profile, "inventory_sheet", "tmp3F5D")
    has_inv = any(inv_sheet_name.lower() in name.lower() for name in wb.sheetnames)
    if has_inv:
        try:
            inv_df, _ = parse_inventory_sheet(wb, profile)
            if not inv_df.empty and "item_name" in inv_df.columns and "rate_per_unit" in inv_df.columns:
                price_list = pd.DataFrame({
                    "sku": inv_df["item_name"],
                    "distributor_price": pd.to_numeric(inv_df["rate_per_unit"], errors="coerce"),
                    "sub_distributor_price": pd.to_numeric(inv_df["rate_per_unit"], errors="coerce"),
                    "retail_price": pd.to_numeric(inv_df["rate_per_unit"], errors="coerce"),
                }).dropna(subset=["sku"]).reset_index(drop=True)
                return price_list
        except Exception:
            pass

    # 7. Final fallback: Return empty price list DataFrame
    return pd.DataFrame(columns=["sku", "distributor_price", "sub_distributor_price", "retail_price"])



def match_products(line_items_df: pd.DataFrame, price_list_df: pd.DataFrame,
                    profile: ClientProfile) -> pd.DataFrame:
    """Returns line_items_df with added columns: matched_sku, match_score, match_method,
    distributor_price, sub_distributor_price, retail_price."""

    skus = price_list_df["sku"].astype(str).tolist()
    sku_norm_map = {s.strip().upper(): s for s in skus}
    sku_sizes = {s: extract_pack_size(s) for s in skus}
    sku_bases = {s: extract_base_name(s).upper() for s in skus}

    unique_products = line_items_df["product_raw"].dropna().unique()
    match_results = {}
    manual_overrides_norm = {k.strip().upper(): v for k, v in profile.manual_overrides.items()}

    for product in unique_products:
        p_norm = str(product).strip().upper()

        # 1. Exact match (case/space-insensitive) first — cheapest and safest
        if p_norm in sku_norm_map:
            match_results[product] = (sku_norm_map[p_norm], 100.0, "exact")
            continue

        # 2. Manual overrides for unambiguous client-specific shorthands
        if p_norm in manual_overrides_norm:
            match_results[product] = (manual_overrides_norm[p_norm], 100.0, "manual_override")
            continue

        p_size = extract_pack_size(product)
        p_base = extract_base_name(product).upper()
        p_norm_clean = normalize_text(product).upper()

        # 3. Candidate filtering by pack size
        if p_size is not None:
            candidate_skus = [s for s in skus if sku_sizes[s] == p_size]
            if candidate_skus:
                cand_map = {normalize_text(s).upper(): s for s in candidate_skus}
                best = process.extractOne(p_norm_clean, list(cand_map.keys()), scorer=fuzz.token_sort_ratio)
                best_sku = cand_map[best[0]] if best else None
                best_score = best[1] if best else 0.0

                # If token_sort was slightly below threshold due to omitted container words (e.g. 'Bottle'),
                # evaluate token_set_ratio on base names among the size-filtered candidates
                if best_score < profile.fuzzy_match_threshold:
                    cand_base_map = {sku_bases[s]: s for s in candidate_skus}
                    best_base = process.extractOne(p_base, list(cand_base_map.keys()), scorer=fuzz.token_set_ratio)
                    if best_base and best_base[1] >= profile.fuzzy_match_threshold:
                        best_sku = cand_base_map[best_base[0]]
                        best_score = best_base[1]

                if best_sku and best_score >= profile.fuzzy_match_threshold:
                    match_results[product] = (best_sku, float(best_score), "fuzzy")
                else:
                    match_results[product] = (None, float(best_score), "unmatched")
            else:
                # No SKU shares the pack size -> fallback to fuzzy match on base name only, flag fuzzy_no_size_match
                sku_base_map = {sku_bases[s]: s for s in skus}
                best_base = process.extractOne(p_base, list(sku_base_map.keys()), scorer=fuzz.token_sort_ratio)
                if best_base and best_base[1] >= profile.fuzzy_match_threshold:
                    match_results[product] = (sku_base_map[best_base[0]], float(best_base[1]), "fuzzy_no_size_match")
                else:
                    best_set = process.extractOne(p_base, list(sku_base_map.keys()), scorer=fuzz.token_set_ratio)
                    if best_set and best_set[1] >= profile.fuzzy_match_threshold:
                        match_results[product] = (sku_base_map[best_set[0]], float(best_set[1]), "fuzzy_no_size_match")
                    else:
                        match_results[product] = (None, float(best_base[1] if best_base else 0.0), "unmatched")
        else:
            # Product has no pack size token
            cand_map = {normalize_text(s).upper(): s for s in skus}
            best = process.extractOne(p_norm_clean, list(cand_map.keys()), scorer=fuzz.token_sort_ratio)
            if best and best[1] >= profile.fuzzy_match_threshold:
                match_results[product] = (cand_map[best[0]], float(best[1]), "fuzzy")
            else:
                match_results[product] = (None, float(best[1] if best else 0.0), "unmatched")

    match_df = pd.DataFrame([
        {"product_raw": k, "matched_sku": v[0], "match_score": v[1], "match_method": v[2]}
        for k, v in match_results.items()
    ])

    out = line_items_df.merge(match_df, on="product_raw", how="left")
    out = out.merge(
        price_list_df.rename(columns={"sku": "matched_sku"}),
        on="matched_sku", how="left"
    )
    return out
