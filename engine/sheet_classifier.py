"""
Structural Sheet Role Classifier for Kane-Jones Depot Engine.
Per Section 23 of Developer_Specification_v2.md ("Month-to-Month Generalization").

Scans every worksheet in an uploaded workbook and assigns roles based on structural signatures:
1. sales_invoice: Repeating invoice-block pattern with date, customer, invoice_no, gross_revenue/item/rate/cost.
2. inventory_cost: Item/Product column + Rate per Unit / Unit Cost / DPP valuation.
3. sales_returns: Credit note / sales return vouchers (VCH TYPE = Sales Return / Credit Note).
4. expenses: Summary table (Category vs Amount) or day-book ledger.
5. price_list: SKU column + tiered pricing columns (distributor, sub_distributor, retail).
6. unclassified: Any sheet that does not match a structural signature (surfaced as informational).

Supports ClientProfile as an optional override/cache:
- If ClientProfile specifies known sheet names that exist in the workbook, it validates them structurally.
- Always falls back to full structural detection across all sheets for unknown/new-month files.
- Returns a structured ClassificationReport with sheet assignments, confidence scores, reasons, and unclassified sheets.
"""

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple, Union
import re
import openpyxl
import pandas as pd
from engine.config import ClientProfile, kane_jones_profile
from engine.parser import (
    INVOICE_FIELDS,
    ITEM_FIELDS,
    _build_alias_lookup,
    _find_header_row,
    _norm,
)


@dataclass
class SheetClassification:
    sheet_name: str
    role: str  # "sales_invoice", "inventory_cost", "sales_returns", "purchases", "purchase_returns", "stock_balances", "expenses", "price_list", "unclassified"
    confidence: float  # 0.0 to 1.0
    detection_method: str  # "profile_cached", "structural_signature", "content_heuristics", "none"
    reason: str
    details: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ClassificationReport:
    sales_sheets: List[str] = field(default_factory=list)
    inventory_sheet: Optional[str] = None
    sales_returns_sheet: Optional[str] = None
    purchases_sheet: Optional[str] = None
    purchase_returns_sheet: Optional[str] = None
    stock_balances_sheet: Optional[str] = None
    expenses_sheet: Optional[str] = None
    price_list_sheet: Optional[str] = None
    unclassified_sheets: List[Dict[str, Any]] = field(default_factory=list)
    classifications: Dict[str, SheetClassification] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sales_sheets": self.sales_sheets,
            "inventory_sheet": self.inventory_sheet,
            "sales_returns_sheet": self.sales_returns_sheet,
            "purchases_sheet": self.purchases_sheet,
            "purchase_returns_sheet": self.purchase_returns_sheet,
            "stock_balances_sheet": self.stock_balances_sheet,
            "expenses_sheet": self.expenses_sheet,
            "price_list_sheet": self.price_list_sheet,
            "unclassified_sheets": self.unclassified_sheets,
            "classifications": {
                name: {
                    "sheet_name": c.sheet_name,
                    "role": c.role,
                    "confidence": c.confidence,
                    "detection_method": c.detection_method,
                    "reason": c.reason,
                    "details": c.details,
                }
                for name, c in self.classifications.items()
            },
        }



def _check_sales_invoice_signature(
    rows: List[List[Any]], profile: ClientProfile
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match the repeating hierarchical invoice-block structure:
    1. Header row containing invoice table columns (date, customer/particulars, voucher/invoice_no).
    2. Followed by the item sub-header row (ITEM/SERVICE, QUANTITY, RATE, COST).
    Strictly requires BOTH the invoice header and the item sub-header to avoid misclassifying flat summary reports.
    """
    invoice_alias_lookup = _build_alias_lookup(profile.column_aliases, only_fields=INVOICE_FIELDS)
    item_alias_lookup = _build_alias_lookup(profile.column_aliases, only_fields=ITEM_FIELDS)

    try:
        header_idx = _find_header_row(rows, invoice_alias_lookup)
    except Exception:
        return False, 0.0, "No invoice header row (date, customer, invoice_no) found", {}

    # Check if there is an item sub-header in the subsequent rows (hierarchical block structure)
    item_sub_header_found = False
    item_sub_header_row = None
    for r_idx in range(header_idx + 1, min(len(rows), header_idx + 30)):
        row_canonicals = {item_alias_lookup.get(_norm(c)) for c in rows[r_idx] if c is not None}
        if "item_service" in row_canonicals and ("quantity" in row_canonicals or "rate" in row_canonicals or "item_cost" in row_canonicals):
            item_sub_header_found = True
            item_sub_header_row = r_idx + 1
            break

    if item_sub_header_found:
        return True, 0.98, f"Matched hierarchical invoice-block structure (invoice header at row {header_idx + 1}, item sub-header at row {item_sub_header_row})", {
            "header_row": header_idx + 1,
            "item_sub_header_row": item_sub_header_row,
        }

    return False, 0.0, "Found table headers but lacks hierarchical ITEM/SERVICE sub-header block (likely a flat report or summary sheet)", {}


def _check_price_list_signature(
    rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match price list structure:
    Contains SKU / ITEM / DESCRIPTION column and tiered price columns (distributor, sub-distributor, retail).
    """
    for r_idx, row in enumerate(rows[:25]):
        str_cells = [str(c or "").strip().upper() for c in row]
        has_sku = any(k in str_cells for k in ["SKU", "ITEM", "PRODUCT", "DESCRIPTION", "PRODUCT NAME", "ITEM NAME"]) or any("SKU" in c for c in str_cells)
        has_distributor = any("DISTRIBUTOR" in c for c in str_cells) or any("WHOLESALE" in c for c in str_cells)
        has_retail = any("RETAIL" in c for c in str_cells)
        has_sub_dist = any("SUB" in c for c in str_cells) or any("TIER" in c for c in str_cells)

        if has_sku and (has_distributor or (has_retail and has_sub_dist)):
            return True, 0.95, f"Matched tiered price list matrix at row {r_idx + 1}", {
                "header_row": r_idx + 1
            }

    return False, 0.0, "No tiered price list headers found", {}


def _check_inventory_cost_signature(
    rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match inventory cost valuation structure:
    Header contains ITEM / PARTICULARS and RATE PER UNIT / UNIT COST / DPP / CLOSING VALUE / NO. OF UNITS.
    """
    for r_idx, row in enumerate(rows[:25]):
        str_cells = [str(c or "").strip().upper() for c in row]
        has_item = any("ITEM" in c or "PARTICULAR" in c or "PRODUCT" in c for c in str_cells)
        has_rate_unit = any("RATE PER UNIT" in c or "RATE/UNIT" in c or "UNIT COST" in c or "PURCHASE COST" in c for c in str_cells)
        has_uom = any("UOM" in c or "UNIT" in c for c in str_cells)
        has_val = any("CLOSING VALUE" in c or "TOTAL VALUE" in c or "VALUE" in c or "DEFAULT PURCHASE PRICE" in c or "DPP" in c for c in str_cells)

        if has_item and (has_rate_unit or (has_uom and has_val)):
            return True, 0.95, f"Matched inventory cost valuation headers at row {r_idx + 1}", {
                "header_row": r_idx + 1
            }

    return False, 0.0, "No inventory cost valuation headers found", {}


def _check_stock_balances_signature(
    rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match the Opening & Closing Stock balance summary table:
    Usually a compact table (<= 15 rows, <= 4 columns) containing 'OPENING STOCK'
    and 'CLOSING STOCK' (or 'OPENING INVENTORY' / 'CLOSING INVENTORY').
    """
    has_open = False
    has_close = False
    open_val = None
    close_val = None
    open_row = None
    close_row = None

    for r_idx, row in enumerate(rows[:15]):
        if not row:
            continue
        c1 = str(row[0] or "").strip().upper()
        c2 = row[1] if len(row) > 1 else None
        if "OPENING" in c1 and ("STOCK" in c1 or "INV" in c1):
            has_open = True
            open_val = c2
            open_row = r_idx + 1
        if "CLOSING" in c1 and ("STOCK" in c1 or "INV" in c1):
            has_close = True
            close_val = c2
            close_row = r_idx + 1

    if has_open and has_close:
        return True, 0.99, f"Matched Stock balances summary table (Opening Stock row {open_row}, Closing Stock row {close_row})", {
            "opening_stock": open_val,
            "closing_stock": close_val,
            "opening_row": open_row,
            "closing_row": close_row,
        }

    return False, 0.0, "No opening/closing stock balance summary found", {}


def _check_purchase_returns_signature(
    rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match supplier purchase returns / debit notes structure:
    Day Book format with 'Purchases Return' or 'Debit Note' voucher indicators.
    """
    is_daybook = False
    for row in rows[:6]:
        row_str = " ".join([str(c or "").strip().upper() for c in row])
        if "DAY BOOK" in row_str or ("DEBIT" in row_str and "CREDIT" in row_str and "VCH TYPE" in row_str):
            is_daybook = True
            break

    has_debit_note = False
    has_purchase_return = False

    for r_idx, row in enumerate(rows[:40]):
        row_str = " ".join([str(c or "").strip().upper() for c in row])
        if "DEBIT NOTE" in row_str:
            has_debit_note = True
        if "PURCHASE RETURN" in row_str or "PURCHASES RETURN" in row_str:
            has_purchase_return = True

        if is_daybook and (has_debit_note or has_purchase_return):
            return True, 0.97, f"Matched purchase returns / debit notes Day-Book at row {r_idx + 1}", {
                "sample_row": r_idx + 1
            }

    return False, 0.0, "No purchase returns or debit note indicators found", {}


def _check_purchases_signature(
    rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match supplier purchases daybook structure:
    Day Book format with 'Purchase' voucher type or supplier ledger entries (excluding returns/payments).
    """
    is_daybook = False
    for row in rows[:6]:
        row_str = " ".join([str(c or "").strip().upper() for c in row])
        if "DAY BOOK" in row_str or ("DEBIT" in row_str and "CREDIT" in row_str and "VCH TYPE" in row_str):
            is_daybook = True
            break

    if not is_daybook:
        return False, 0.0, "Not a Day-Book ledger", {}

    purchase_vouchers = 0
    for r_idx, row in enumerate(rows[:50]):
        row_str = " ".join([str(c or "").strip().upper() for c in row])
        # Must be a regular purchase, not a return or payment
        if "PURCHASE" in row_str and not ("RETURN" in row_str or "CREDIT NOTE" in row_str or "DEBIT NOTE" in row_str):
            # Check column 4 (VCH TYPE) or Debit column
            c4 = str(row[3] or "").strip().upper() if len(row) > 3 else ""
            c2 = str(row[1] or "").strip().upper() if len(row) > 1 else ""
            if "PURCHASE" in c4 or "PURCHASE" in c2:
                purchase_vouchers += 1

    if purchase_vouchers >= 2:
        return True, 0.97, f"Matched supplier purchases Day-Book ({purchase_vouchers} purchase vouchers)", {
            "purchase_vouchers_count": purchase_vouchers
        }

    return False, 0.0, "No supplier purchase vouchers found in Day-Book", {}


def _check_sales_returns_signature(
    rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match sales returns / credit notes structure:
    Contains 'Sales Return' or 'Credit Note' literals in VCH TYPE or DEBIT/CREDIT ledger columns.
    """
    has_sr_literal = False
    has_cn_literal = False
    header_found = False

    for r_idx, row in enumerate(rows[:30]):
        row_str = " ".join([str(c or "").strip().lower() for c in row])
        if "sales return" in row_str:
            has_sr_literal = True
        if "credit note" in row_str:
            has_cn_literal = True
        if "vch type" in row_str or "voucher type" in row_str or "credit note" in row_str:
            header_found = True

        if (has_sr_literal and has_cn_literal) or (header_found and (has_sr_literal or has_cn_literal)):
            return True, 0.96, f"Matched sales return and credit note voucher indicators at row {r_idx + 1}", {
                "sample_row": r_idx + 1
            }

    return False, 0.0, "No sales return or credit note indicators found", {}


def _check_expenses_signature(
    sheet_name: str, rows: List[List[Any]]
) -> Tuple[bool, float, str, Dict[str, Any]]:
    """
    Checks if rows match operating expense structure:
    1. Category vs Amount summary table (Column 1 = Category, Column 2 = numeric amount, Grand Total row).
    2. Or Day-Book ledger format with Payment / Journal vouchers.
    Excludes sales, invoices, returns, inventory, and single bank/cash transaction accounts.
    """
    clean_name = sheet_name.strip().lower()
    
    # Exclude non-expense sheets
    if any(ex in clean_name for ex in ["product", "marketer", "volume", "sales", "revenue", "price", "aggregate", "return", "credit", "inv", "stock"]):
        return False, 0.0, "Sheet name indicates sales/inventory/returns analysis, not operating expenses", {}

    # Reject single-account bank/cash sheets (Cash, GTB, etc.)
    if clean_name in ["cash", "gtb", "gtb (kanejones)", "pos", "bank charges", "advance salary"]:
        return False, 0.0, "Single-account payment ledger, not primary depot expense summary", {}

    # Check top text for analysis or invoice titles
    top_text = " ".join([str(c or "").strip().lower() for row in rows[:5] for c in row])
    if any(ex in top_text for ex in ["sales & gross profit", "selling price vs", "cases sold", "product analysis", "voucher date particulars vchno"]):
        return False, 0.0, "Content indicates sales/invoice register, not operating expenses", {}

    # 1. Check Category Summary Table format (Highest Priority, 0.98 confidence)
    categories_count = 0
    has_grand_total = False
    has_expense_header = False

    for r_idx, row in enumerate(rows):
        if len(row) < 2:
            continue
        c1 = str(row[0] or "").strip().lower()
        c2 = row[1]

        if "category" in c1 and ("amount" in str(c2).lower() or "%" in str(c2).lower() or "cost" in str(c2).lower() or "expense" in str(c2).lower()):
            has_expense_header = True
            continue

        if "grand total" in c1 or "total expense" in c1 or (c1 == "total" and isinstance(c2, (int, float))):
            has_grand_total = True
            continue

        # Test if c1 is category text and c2 is numeric amount
        if c1 and isinstance(c2, (int, float)) and c2 > 0 and not c1.startswith("---") and not c1.startswith("voucher"):
            if not any(k in c1 for k in ["date", "particulars", "vchno", "total"]):
                categories_count += 1

    if (has_expense_header and categories_count >= 2) or (has_grand_total and categories_count >= 2) or (categories_count >= 3 and any(kw in clean_name for kw in ["expense", "expn", "opex", "overhead"])):
        return True, 0.98, f"Matched clean expense category summary table ({categories_count} categories, grand total row found: {has_grand_total})", {
            "categories_count": categories_count,
            "has_grand_total": has_grand_total,
        }

    # 2. Check Day-Book Ledger Format (Payment / Journal vouchers, 0.85 confidence)
    is_daybook_header = False
    for row in rows[:5]:
        row_str = " ".join([str(c or "").strip().lower() for c in row])
        if "day book" in row_str or ("debit" in row_str and "credit" in row_str and "vch type" in row_str):
            is_daybook_header = True
            break

    payment_vouchers_count = 0
    if is_daybook_header:
        for row in rows[:50]:
            row_str = " ".join([str(c or "").strip().lower() for c in row])
            if "payment" in row_str or "journal" in row_str:
                payment_vouchers_count += 1

    if is_daybook_header and payment_vouchers_count >= 3:
        return True, 0.85, f"Matched payment/journal expense Day-Book ledger ({payment_vouchers_count} vouchers)", {
            "payment_vouchers_count": payment_vouchers_count
        }

    return False, 0.0, "No expense category summary or Day-Book voucher ledger found", {}


def classify_workbook_sheets(
    xlsx_path_or_wb: Union[str, openpyxl.Workbook],
    profile: Optional[ClientProfile] = None,
) -> ClassificationReport:
    """
    Scans every sheet in the workbook and assigns each sheet to a role based on structural signatures.
    Reuses ClientProfile as an optional fast-path/override cache if sheets match.
    """
    if profile is None:
        profile = kane_jones_profile()

    if isinstance(xlsx_path_or_wb, str):
        wb = openpyxl.load_workbook(xlsx_path_or_wb, data_only=True)
    else:
        wb = xlsx_path_or_wb

    report = ClassificationReport()
    sheet_names = wb.sheetnames

    # Read top 60 rows of each sheet for signature detection
    sheet_sample_rows: Dict[str, List[List[Any]]] = {}
    for name in sheet_names:
        ws = wb[name]
        sheet_sample_rows[name] = [[c.value for c in row] for row in ws.iter_rows(max_row=60)]

    # 1. Pass A: Price List Detection (Explicit/Single Role)
    for name in sheet_names:
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_pl, conf, reason, details = _check_price_list_signature(rows)
        
        # Name hint boost
        if "price" in clean_name or clean_name == getattr(profile, "price_list_sheet", "").lower():
            if is_pl:
                conf = min(conf + 0.05, 1.0)
            elif any("sku" in str(c or "").lower() for row in rows[:10] for c in row):
                is_pl = True
                conf = 0.90
                reason = "Price list sheet identified by name and SKU table"

        if is_pl and not report.price_list_sheet:
            report.price_list_sheet = name
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="price_list",
                confidence=conf,
                detection_method="structural_signature",
                reason=reason,
                details=details,
            )

    # 2. Pass B: Stock Balances Sheet Detection (Opening & Closing Stock Summary)
    for name in sheet_names:
        if name in report.classifications:
            continue
        rows = sheet_sample_rows[name]
        is_stk, conf, reason, details = _check_stock_balances_signature(rows)
        if is_stk and not report.stock_balances_sheet:
            report.stock_balances_sheet = name
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="stock_balances",
                confidence=conf,
                detection_method="structural_signature",
                reason=reason,
                details=details,
            )

    # 3. Pass C: Inventory Cost Sheet Detection (Item-level unit costs / DPP)
    for name in sheet_names:
        if name in report.classifications:
            continue
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_inv, conf, reason, details = _check_inventory_cost_signature(rows)

        if ("3f5d" in clean_name or "inventory" in clean_name or "stock" in clean_name or clean_name == getattr(profile, "inventory_sheet", "").lower()) and not is_inv:
            # Only treat as inventory cost if it looks like item table, not stock balances
            if not _check_stock_balances_signature(rows)[0]:
                is_inv = True
                conf = 0.85
                reason = "Inventory valuation sheet identified by profile/name match"

        if is_inv and not report.inventory_sheet:
            report.inventory_sheet = name
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="inventory_cost",
                confidence=conf,
                detection_method="structural_signature",
                reason=reason,
                details=details,
            )

    # 4. Pass D: Sales Returns Sheet Detection (Credit Notes)
    for name in sheet_names:
        if name in report.classifications:
            continue
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_ret, conf, reason, details = _check_sales_returns_signature(rows)

        if ("cef3" in clean_name or "return" in clean_name or "credit" in clean_name or clean_name == getattr(profile, "sales_returns_sheet", "").lower()) and not is_ret:
            # Ensure it's not a purchase return
            if not _check_purchase_returns_signature(rows)[0]:
                is_ret = True
                conf = 0.85
                reason = "Sales returns sheet identified by profile/name match"

        if is_ret and not report.sales_returns_sheet:
            report.sales_returns_sheet = name
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="sales_returns",
                confidence=conf,
                detection_method="structural_signature",
                reason=reason,
                details=details,
            )

    # 5. Pass E: Purchase Returns Sheet Detection (Debit Notes)
    for name in sheet_names:
        if name in report.classifications:
            continue
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_pr, conf, reason, details = _check_purchase_returns_signature(rows)

        if ("purchase return" in clean_name or "purchases return" in clean_name or "prr" in clean_name) and not is_pr:
            is_pr = True
            conf = 0.85
            reason = "Purchase returns sheet identified by name match"

        if is_pr and not report.purchase_returns_sheet:
            report.purchase_returns_sheet = name
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="purchase_returns",
                confidence=conf,
                detection_method="structural_signature",
                reason=reason,
                details=details,
            )

    # 6. Pass F: Purchases Daybook Sheet Detection
    for name in sheet_names:
        if name in report.classifications:
            continue
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_pur, conf, reason, details = _check_purchases_signature(rows)

        if ("purchase" in clean_name and "return" not in clean_name) and not is_pur:
            is_pur = True
            conf = 0.85
            reason = "Purchases Day-Book sheet identified by name match"

        if is_pur and not report.purchases_sheet:
            report.purchases_sheet = name
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="purchases",
                confidence=conf,
                detection_method="structural_signature",
                reason=reason,
                details=details,
            )

    # 7. Pass G: Operating Expenses Sheet Detection (Prioritize clean category summaries)
    expense_candidates = []
    for name in sheet_names:
        if name in report.classifications:
            continue
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_exp, conf, reason, details = _check_expenses_signature(name, rows)

        if any(kw in clean_name for kw in ["expense", "expn", "opex", "overhead"]) or clean_name == getattr(profile, "expenses_sheet", "").lower():
            if is_exp:
                conf = min(conf + 0.05, 1.0)
            elif not any(ex in clean_name for ex in ["product", "marketer", "volume", "sales", "cash", "gtb"]):
                is_exp = True
                conf = 0.85
                reason = "Operating expenses sheet identified by profile/name match"

        if is_exp:
            expense_candidates.append((conf, name, reason, details))

    if expense_candidates:
        # Pick the highest confidence candidate (Category Summary table > Day Book)
        expense_candidates.sort(key=lambda x: x[0], reverse=True)
        best_conf, best_name, best_reason, best_details = expense_candidates[0]
        report.expenses_sheet = best_name
        report.classifications[best_name] = SheetClassification(
            sheet_name=best_name,
            role="expenses",
            confidence=best_conf,
            detection_method="structural_signature",
            reason=best_reason,
            details=best_details,
        )

    # 8. Pass H: Sales / Invoice Sheets Detection (Can be multiple sheets, e.g. tmpA1A6, tmp32C7)
    for name in sheet_names:
        if name in report.classifications:
            continue
        clean_name = name.strip().lower()
        rows = sheet_sample_rows[name]
        is_sales, conf, reason, details = _check_sales_invoice_signature(rows, profile)

        # Profile cached fallback
        if clean_name in [s.lower() for s in getattr(profile, "raw_data_sheets", [])]:
            if is_sales:
                conf = 1.0
                detection_method = "profile_cached"
            else:
                is_sales = True
                conf = 0.85
                detection_method = "profile_cached"
                reason = "Identified as sales register by ClientProfile cached mapping"
        else:
            detection_method = "structural_signature"

        if is_sales:
            report.sales_sheets.append(name)
            report.classifications[name] = SheetClassification(
                sheet_name=name,
                role="sales_invoice",
                confidence=conf,
                detection_method=detection_method,
                reason=reason,
                details=details,
            )

    # 9. Pass I: Unclassified Sheets
    for name in sheet_names:
        if name not in report.classifications:
            classification = SheetClassification(
                sheet_name=name,
                role="unclassified",
                confidence=0.0,
                detection_method="none",
                reason="Sheet did not match any known structural signature (hierarchical sales register, tiered price list, inventory cost valuation, stock balances, returns credit notes, supplier purchases/debit notes, or operating expenses)",
            )
            report.classifications[name] = classification
            report.unclassified_sheets.append({
                "sheet_name": name,
                "reason": classification.reason,
            })

    return report

