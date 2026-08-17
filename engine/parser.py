"""
Parses the raw invoice-block sheet(s) (Kane-Jones calls them tmpA1A6 / tmp32C7)
into two clean, normalized pandas DataFrames: invoices and line_items.

Why not just pandas.read_excel(): the source is NOT a flat table. It's a repeating
block structure per invoice:

    <invoice header row>              date, customer, invoice_no, transaction_value, ...
    <item sub-header row>             "ITEM/SERVICE", "QUANTITY", "RATE", "COST", "DESCRIPTION"
    <item row>                        product name, qty, rate, cost
    <item row>
    ...
    <blank separator row>

This state machine walks the raw grid and reconstructs that structure. It's built to
be config-driven (via ClientProfile.column_aliases) rather than hardcoded to exact
header text, so a client whose sheet says "TRANS VALUE" instead of "TRANSACTION VALUE"
still works. Rows that don't fit the expected pattern (e.g. stray subtotal rows from a
spreadsheet formula bug) are collected as `anomalies` instead of crashing the parse —
you get a report of what was skipped, not a silent wrong number or a stack trace.
"""

import datetime
import re
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
import openpyxl
from engine.config import ClientProfile


def _parse_quantity(val) -> Tuple[Optional[float], Optional[str]]:
    """Parse quantity value. Returns (parsed_qty, error_reason).
    Handles numbers and text quantities like '5 cans', '1,240 pets'.
    If string contains no extractable leading number, returns (None, error_msg).
    """
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return 0.0, None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (pd.isna(val) or pd.isnull(val)):
            return 0.0, None
        return float(val), None
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "")
        if not cleaned:
            return 0.0, None
        m = re.search(r"^\s*([+-]?\d+(?:\.\d+)?)", cleaned)
        if m:
            try:
                return float(m.group(1)), None
            except ValueError:
                pass
        return None, f"Could not extract numeric quantity from text: '{val}'"
    return None, f"Unsupported quantity type '{type(val).__name__}': {val}"


def _parse_float(val, default: float = 0.0) -> float:
    """Safely converts cell values (including comma-formatted strings) to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val) if not (isinstance(val, float) and (pd.isna(val) or pd.isnull(val))) else default
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "")
        try:
            return float(cleaned)
        except ValueError:
            m = re.search(r"([+-]?\d+(?:\.\d+)?)", cleaned)
            if m:
                try:
                    return float(m.group(1))
                except ValueError:
                    pass
    return default


def _norm(s):
    if s is None:
        return ""
    return str(s).strip().upper()


# The invoice header row and the per-block item sub-header row are DIFFERENT tables
# that happen to reuse the literal header text "COST" for different meanings
# (invoice-level total cost vs. per-line-item cost). They must be resolved with
# separate lookups, built only from the canonical fields relevant to each context —
# otherwise "COST" collapses to whichever canonical happens to iterate last.
INVOICE_FIELDS = {
    "date", "customer", "invoice_no", "po_no", "transaction_value",
    "received_amount", "outstanding_amount", "gross_revenue", "invoice_cost",
    "gross_profit", "pct_profit", "narration",
}
ITEM_FIELDS = {"item_service", "quantity", "rate", "item_cost", "description"}


def _build_alias_lookup(column_aliases: dict, only_fields: set = None) -> dict:
    """Flatten {canonical: [alias, ...]} into {ALIAS_TEXT: canonical} for exact matching.
    If only_fields is given, restrict to those canonical fields (avoids cross-context collisions,
    e.g. 'COST' meaning different things in the invoice row vs. the item sub-header row)."""
    lookup = {}
    for canonical, aliases in column_aliases.items():
        if only_fields is not None and canonical not in only_fields:
            continue
        for a in aliases:
            lookup[_norm(a)] = canonical
    return lookup


def _find_header_row(rows, alias_lookup, required=("date", "customer", "invoice_no")):
    """Find the row index that looks like the main invoice-table header."""
    for i, row in enumerate(rows):
        found = {alias_lookup.get(_norm(c)) for c in row}
        if all(r in found for r in required):
            return i
    raise ValueError(
        "Could not find the main header row (expected columns like "
        "VOUCHER DATE / PARTICULARS / VCHNO). This sheet's structure may differ "
        "from what this client profile expects — check column_aliases in the profile."
    )


def _col_map_from_header(header_row, alias_lookup):
    col_map = {}
    for idx, cell in enumerate(header_row):
        canonical = alias_lookup.get(_norm(cell))
        if canonical:
            col_map[canonical] = idx
    return col_map


def _is_blank_row(row):
    return all(c is None or (isinstance(c, str) and c.strip() == "") for c in row)


def _looks_like_date(v):
    return isinstance(v, (datetime.datetime, datetime.date))


def parse_raw_sheet(ws, profile: ClientProfile, source_tab: str):
    """Parse a single raw worksheet. Returns (invoices: list[dict], line_items: list[dict], anomalies: list[dict])."""
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    invoice_alias_lookup = _build_alias_lookup(profile.column_aliases, only_fields=INVOICE_FIELDS)
    item_alias_lookup = _build_alias_lookup(profile.column_aliases, only_fields=ITEM_FIELDS)

    header_idx = _find_header_row(rows, invoice_alias_lookup)
    col_map = _col_map_from_header(rows[header_idx], invoice_alias_lookup)

    def get(row, canonical):
        idx = col_map.get(canonical)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    invoices = []
    line_items = []
    anomalies = []

    state = "SEEK_INVOICE"
    current_invoice = None
    item_col_map = None  # re-derived per block from the item sub-header row

    for i in range(header_idx + 1, len(rows)):
        row = rows[i]

        if _is_blank_row(row):
            if state == "READING_ITEMS":
                state = "SEEK_INVOICE"
                current_invoice = None
                item_col_map = None
            continue

        if state == "SEEK_INVOICE":
            date_val = get(row, "date")
            invoice_no = get(row, "invoice_no")
            if _looks_like_date(date_val) and invoice_no:
                current_invoice = {
                    "source_tab": source_tab,
                    "row": i + 1,
                    "date": date_val,
                    "customer": get(row, "customer"),
                    "invoice_no": invoice_no,
                    "po_no": get(row, "po_no"),
                    "transaction_value": _parse_float(get(row, "transaction_value")),
                    "received_amount": _parse_float(get(row, "received_amount")),
                    "outstanding_amount": _parse_float(get(row, "outstanding_amount")),
                    "gross_revenue": _parse_float(get(row, "gross_revenue")),
                    "invoice_cost": _parse_float(get(row, "invoice_cost")),
                    "gross_profit": _parse_float(get(row, "gross_profit")),
                    "pct_profit": get(row, "pct_profit"),
                    "narration": get(row, "narration"),
                }
                invoices.append(current_invoice)
                state = "SEEK_ITEM_HEADER"
            else:
                anomalies.append({"row": i + 1, "source_tab": source_tab, "raw": row,
                                   "reason": "Row outside a known block: no date+invoice_no, not blank."})
            continue

        if state == "SEEK_ITEM_HEADER":
            item_header_canonical = {item_alias_lookup.get(_norm(c)) for c in row}
            if "item_service" in item_header_canonical:
                item_col_map = _col_map_from_header(row, item_alias_lookup)
                state = "READING_ITEMS"
            else:
                anomalies.append({"row": i + 1, "source_tab": source_tab, "raw": row,
                                   "reason": "Expected an ITEM/SERVICE sub-header row after invoice header, didn't find one."})
                state = "SEEK_INVOICE"
                current_invoice = None
            continue

        if state == "READING_ITEMS":
            def iget(canonical):
                idx = item_col_map.get(canonical)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            product = iget("item_service")
            if product is None or (isinstance(product, str) and product.strip() == ""):
                # A non-blank row with no product name mid-block is not a valid item line —
                # most commonly a stray subtotal row from a source-workbook formula bug
                # (see report Section 7: "self-referencing SUM ranges"). Treat it the same
                # as the missing blank separator it should have been, so the NEXT row is
                # correctly evaluated as a fresh invoice header rather than misread as an
                # item row (which would silently invent fake products).
                anomalies.append({"row": i + 1, "source_tab": source_tab, "raw": row,
                                   "reason": "Non-blank row mid-item-block with no product name "
                                             "(likely a stray subtotal/formula-bug row) — "
                                             "treated as end of block."})
                state = "SEEK_INVOICE"
                current_invoice = None
                item_col_map = None
                continue

            raw_qty = iget("quantity")
            qty_val, qty_err = _parse_quantity(raw_qty)
            if qty_val is None:
                anomalies.append({
                    "row": i + 1,
                    "source_tab": source_tab,
                    "raw": row,
                    "reason": f"Unparseable line item quantity '{raw_qty}': {qty_err}",
                })
                continue

            parsed_cost = _parse_float(iget("item_cost"))
            if parsed_cost is not None and parsed_cost < 0:
                anomalies.append({
                    "row": i + 1,
                    "source_tab": source_tab,
                    "raw": row,
                    "reason": f"Negative item cost ({parsed_cost}) detected for '{product}' in invoice '{current_invoice['invoice_no']}'. Flagged as data entry anomaly.",
                })

            line_items.append({
                "source_tab": source_tab,
                "row": i + 1,
                "invoice_no": current_invoice["invoice_no"],
                "date": current_invoice["date"],
                "customer": current_invoice["customer"],
                "product_raw": str(product).strip(),
                "quantity": qty_val,
                "rate": _parse_float(iget("rate")),
                "cost": parsed_cost,
                "description": iget("description"),
            })
            continue

    return invoices, line_items, anomalies


def parse_workbook(xlsx_path: str, profile: ClientProfile, classification_report: Optional[Any] = None):
    """Parse raw sales data sheets from workbook. Returns (invoices_df, line_items_df, anomalies_df)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1. Determine sheets to parse
    sheets_to_parse = []
    
    # Check provided classification report
    if classification_report is not None:
        raw_rep_sheets = getattr(classification_report, "sales_sheets", [])
        if isinstance(raw_rep_sheets, list):
            sheets_to_parse = [s for s in raw_rep_sheets if s in wb.sheetnames]

    # Fast path from profile
    if not sheets_to_parse:
        for target in profile.raw_data_sheets:
            if target in wb.sheetnames:
                sheets_to_parse.append(target)
            else:
                clean_target = re.sub(r'(?i)\.xlsx?$', '', target).strip().upper()
                for name in wb.sheetnames:
                    clean_name = re.sub(r'(?i)\.xlsx?$', '', name).strip().upper()
                    if clean_name == clean_target and name not in sheets_to_parse:
                        sheets_to_parse.append(name)
                        break

    # Dynamic structural classification fallback
    if not sheets_to_parse:
        from engine.sheet_classifier import classify_workbook_sheets
        rep = classify_workbook_sheets(wb, profile)
        sheets_to_parse = [s for s in rep.sales_sheets if s in wb.sheetnames]

    if not sheets_to_parse:
        raise ValueError(
            f"No sales register sheets found in workbook. Expected sheets like {profile.raw_data_sheets} or containing invoice headers. Available sheets: {wb.sheetnames}"
        )

    all_invoices, all_line_items, all_anomalies = [], [], []
    for sheet_name in sheets_to_parse:
        ws = wb[sheet_name]
        try:
            invoices, line_items, anomalies = parse_raw_sheet(ws, profile, sheet_name)
            all_invoices.extend(invoices)
            all_line_items.extend(line_items)
            all_anomalies.extend(anomalies)
        except Exception as e:
            all_anomalies.append({
                "row": None,
                "source_tab": sheet_name,
                "reason": f"Failed parsing sheet '{sheet_name}': {str(e)}",
            })

    invoices_df = pd.DataFrame(all_invoices)
    line_items_df = pd.DataFrame(all_line_items)
    anomalies_df = pd.DataFrame(all_anomalies)

    if invoices_df.empty:
        raise ValueError(
            f"No invoices could be extracted from sheet(s) {sheets_to_parse}. "
            "Please check that the column headers match the expected sales register format."
        )

    invoices_df["date"] = pd.to_datetime(invoices_df["date"])
    invoices_df["gross_profit"] = pd.to_numeric(invoices_df["gross_profit"], errors="coerce").fillna(0.0)
    invoices_df["gross_revenue"] = pd.to_numeric(invoices_df["gross_revenue"], errors="coerce").fillna(0.0)
    invoices_df["is_loss_making"] = invoices_df["gross_profit"] < 0

    if not line_items_df.empty:
        line_items_df["date"] = pd.to_datetime(line_items_df["date"])
        line_items_df["quantity"] = pd.to_numeric(line_items_df["quantity"], errors="coerce").fillna(0.0)
        line_items_df["rate"] = pd.to_numeric(line_items_df["rate"], errors="coerce").fillna(0.0)
        line_items_df["cost"] = pd.to_numeric(line_items_df["cost"], errors="coerce").fillna(0.0)

    return invoices_df, line_items_df, anomalies_df


def parse_inventory_sheet(xlsx_or_wb, profile: ClientProfile = None, classification_report: Optional[Any] = None):
    """
    Parses the hierarchical inventory/cost sheet into a clean DataFrame of leaf SKU items and extracts anomalies.
    """
    if profile is None:
        from engine.config import kane_jones_profile
        profile = kane_jones_profile()

    if isinstance(xlsx_or_wb, str):
        wb = openpyxl.load_workbook(xlsx_or_wb, data_only=True)
    else:
        wb = xlsx_or_wb

    sheet_name = None
    if classification_report is not None and getattr(classification_report, "inventory_sheet", None):
        if classification_report.inventory_sheet in wb.sheetnames:
            sheet_name = classification_report.inventory_sheet

    if not sheet_name:
        target_name = getattr(profile, "inventory_sheet", "tmp3F5D")
        if target_name in wb.sheetnames:
            sheet_name = target_name
        else:
            from engine.sheet_classifier import classify_workbook_sheets
            rep = classify_workbook_sheets(wb, profile)
            if rep.inventory_sheet and rep.inventory_sheet in wb.sheetnames:
                sheet_name = rep.inventory_sheet

    if not sheet_name:
        return pd.DataFrame(), []

    ws = wb[sheet_name]
    rows = []
    anomalies = []

    # Locate header row containing 'ITEM' and ('UOM' or 'RATE PER UNIT' or 'NO. OF UNITS')
    header_row_idx = None
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
        if "ITEM" in row_vals and ("UOM" in row_vals or "RATE PER UNIT" in row_vals or "NO. OF UNITS" in row_vals or "RATE" in row_vals):
            header_row_idx = r
            break

    start_r = (header_row_idx + 1) if header_row_idx else 5

    for r in range(start_r, ws.max_row + 1):
        item_val = ws.cell(r, 1).value
        units_val = ws.cell(r, 2).value
        uom_val = ws.cell(r, 3).value
        rate_val = ws.cell(r, 4).value
        val_val = ws.cell(r, 5).value
        dpp_val = ws.cell(r, 6).value if ws.max_column >= 6 else None

        if not item_val or str(item_val).strip() == "":
            continue

        item_str = str(item_val).strip()
        if "TOTAL" in item_str.upper() or item_str.startswith("---"):
            continue

        units = _parse_float(units_val)
        rate = _parse_float(rate_val)
        val = _parse_float(val_val)
        dpp = _parse_float(dpp_val) if dpp_val is not None else None
        uom_str = str(uom_val).strip() if uom_val else "Ctn"

        if units <= 0:
            anomalies.append({
                "row": r,
                "source_tab": sheet_name,
                "type": "negative_or_zero_inventory_units",
                "item": item_str,
                "units": units,
                "reason": f"Negative or zero inventory units in {sheet_name}: units={units}"
            })

        if rate <= 0:
            anomalies.append({
                "row": r,
                "source_tab": sheet_name,
                "type": "zero_or_negative_inventory_rate",
                "item": item_str,
                "rate": rate,
                "reason": f"Negative or zero inventory cost in {sheet_name}: rate={rate}"
            })

        calc_val = units * rate
        if abs(val - calc_val) > max(100.0, abs(val) * 0.02) and units > 0 and rate > 0:
            anomalies.append({
                "row": r,
                "source_tab": sheet_name,
                "type": "value_mismatch",
                "item": item_str,
                "units": units,
                "rate": rate,
                "expected_value": calc_val,
                "actual_value": val,
                "diff": val - calc_val,
                "reason": f"Inventory value mismatch: reported={val}, calculated (units*rate)={calc_val}"
            })

        rows.append({
            "item_name": item_str,
            "units": units,
            "uom": uom_str,
            "rate_per_unit": rate,
            "value": val,
            "default_purchase_price": dpp,
            "source_row": r,
            "source_tab": sheet_name,
        })

    inv_df = pd.DataFrame(rows)
    return inv_df, anomalies


def parse_sales_returns_sheet(xlsx_or_wb, profile: ClientProfile = None, classification_report: Optional[Any] = None):
    """
    Parses the sales returns / credit notes sheet into a clean DataFrame
    of return line items with empties classification.
    """
    if profile is None:
        from engine.config import kane_jones_profile
        profile = kane_jones_profile()

    if isinstance(xlsx_or_wb, str):
        wb = openpyxl.load_workbook(xlsx_or_wb, data_only=True)
    else:
        wb = xlsx_or_wb

    sheet_name = None
    if classification_report is not None and getattr(classification_report, "sales_returns_sheet", None):
        if classification_report.sales_returns_sheet in wb.sheetnames:
            sheet_name = classification_report.sales_returns_sheet

    if not sheet_name:
        target_name = getattr(profile, "sales_returns_sheet", "tmpCEF3")
        if target_name in wb.sheetnames:
            sheet_name = target_name
        else:
            from engine.sheet_classifier import classify_workbook_sheets
            rep = classify_workbook_sheets(wb, profile)
            if rep.sales_returns_sheet and rep.sales_returns_sheet in wb.sheetnames:
                sheet_name = rep.sales_returns_sheet

    if not sheet_name:
        return pd.DataFrame(), []

    ws = wb[sheet_name]
    returns = []
    anomalies = []

    current_voucher = None
    state = "SEEKING_HEADER"

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        c6 = ws.cell(r, 6).value
        c7 = ws.cell(r, 7).value

        # Check for footer row
        if c5 and "TOTAL" in str(c5).upper():
            break

        # Check for Credit Note header row
        # Col 2: 'Sales Return', Col 3: Customer, Col 4: 'Credit Note', Col 5: Voucher No, Col 7: Amount
        if c2 and str(c2).strip().lower() == "sales return" and c4 and "credit note" in str(c4).strip().lower():
            vch_date = c1
            if isinstance(vch_date, datetime.datetime):
                date_str = vch_date.strftime("%Y-%m-%d")
            elif isinstance(vch_date, str):
                date_str = vch_date[:10]
            else:
                date_str = str(vch_date)

            customer = str(c3).strip() if c3 else "Unknown"
            vch_no = str(c5).strip() if c5 else ""
            total_amount = _parse_float(c7)

            current_voucher = {
                "date": date_str,
                "customer": customer,
                "voucher_no": vch_no,
                "total_amount": total_amount,
                "items": []
            }
            state = "SEEKING_ITEMS"
            continue

        if state in ["SEEKING_ITEMS", "READING_ITEMS"]:
            if c2 and str(c2).strip().upper() == "ITEM/SERVICE":
                state = "READING_ITEMS"
                continue

            if state == "READING_ITEMS":
                if c2 is not None and str(c2).strip() != "":
                    item_name = str(c2).strip()
                    if item_name.lower().startswith("narration:"):
                        continue

                    raw_qty = c3
                    qty, qty_err = _parse_quantity(raw_qty)
                    rate = _parse_float(c4)
                    desc = str(c5).strip() if c5 else ""

                    if qty_err:
                        anomalies.append({
                            "row": r,
                            "source_tab": sheet_name,
                            "type": "unparseable_return_quantity",
                            "item": item_name,
                            "reason": qty_err
                        })

                    line_val = (qty or 0.0) * rate

                    empties_kws = [k.lower() for k in profile.empties_keywords]
                    is_empties = any(kw in item_name.lower() for kw in empties_kws)
                    item_type = "Empties" if is_empties else "Product"

                    item_record = {
                        "date": current_voucher["date"],
                        "customer": current_voucher["customer"],
                        "voucher_no": current_voucher["voucher_no"],
                        "item_name": item_name,
                        "raw_quantity": str(raw_qty) if raw_qty is not None else "",
                        "quantity": qty or 0.0,
                        "rate": rate,
                        "return_value": line_val,
                        "item_type": item_type,
                        "description": desc,
                        "source_row": r,
                        "source_tab": sheet_name,
                    }
                    current_voucher["items"].append(item_record)
                    returns.append(item_record)
                elif c2 is None or str(c2).strip() == "":
                    state = "SEEKING_HEADER"

    returns_df = pd.DataFrame(returns)
    return returns_df, anomalies
